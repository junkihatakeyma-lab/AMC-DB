import json
import os
from openpyxl import Workbook
import re

# We will read 検索画面.html, extract DATA = {...}, and generate the excel files.
HTML_PATH = "検索画面.html"

def extract_json_from_html():
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        content = f.read()
    
    # Find const DATA = { ... };
    match = re.search(r'const DATA = (\{.*?\});\n', content, re.DOTALL)
    if not match:
        raise ValueError("Could not find DATA in HTML")
    
    json_str = match.group(1)
    return json.loads(json_str)

def create_request_excel(req):
    wb = Workbook()
    ws = wb.active
    
    # Mapping: 品名=F10, 依頼No=Q8, 数量=F13, 生地色=F16, 用途=F19, 納期=F22, 出荷先=F31, 規格=F34+L34, 備考=F42
    ws['Q8'] = req['request_no']
    ws['F10'] = req['hinmei']
    ws['F13'] = req['qty']
    ws['F16'] = req['kiji']
    ws['F19'] = req['yoto']
    ws['F22'] = req['noki_raw']
    ws['F31'] = req['dest']
    
    spec = req.get('spec', '')
    specs = spec.split('\n', 1)
    ws['F34'] = specs[0] if len(specs) > 0 else ""
    ws['L34'] = specs[1] if len(specs) > 1 else ""
    
    ws['F42'] = req['biko']
    
    os.makedirs(os.path.dirname(req['file']), exist_ok=True)
    wb.save(req['file'])

def create_bom_excel(bom):
    wb = Workbook()
    ws = wb.active
    
    if bom['layout_ok']:
        ws['A1'] = "役割"
        ws['B1'] = "部品番号"
        ws['C1'] = "仕様メモ"
        
        for i, comp in enumerate(bom['components']):
            row = i + 2
            ws.cell(row=row, column=1, value=comp.get('role', ''))
            ws.cell(row=row, column=2, value=comp.get('part_no', ''))
            ws.cell(row=row, column=3, value=comp.get('note', ''))
    else:
        # Example of exception layout
        ws['A1'] = "This is a messy layout"
        ws['A2'] = "Something else"
        ws['B3'] = "Part: " + str(bom['components'][0].get('part_no', ''))
        
    os.makedirs(os.path.dirname(bom['file']), exist_ok=True)
    wb.save(bom['file'])
    
def generate_all():
    data = extract_json_from_html()
    
    for req in data.get('requests', []):
        create_request_excel(req)
        
    for bom in data.get('boms', []):
        create_bom_excel(bom)
        
    print("Dummy data generated successfully.")

if __name__ == "__main__":
    generate_all()
