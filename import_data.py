import os
import glob
import sqlite3
import hashlib
import openpyxl
import warnings
import re
import pandas as pd
from datetime import datetime
from database import get_db, init_db

def get_file_hash(filepath):
    hasher = hashlib.md5()
    with open(filepath, 'rb') as f:
        buf = f.read()
        hasher.update(buf)
    return hasher.hexdigest()

def check_file_changed(c, filepath):
    mtime = os.path.getmtime(filepath)
    file_hash = get_file_hash(filepath)
    c.execute("SELECT mtime, hash FROM files WHERE file_path=?", (filepath,))
    row = c.fetchone()
    if row:
        if row['hash'] == file_hash:
            return False # Not changed
    c.execute("INSERT OR REPLACE INTO files (file_path, mtime, hash) VALUES (?, ?, ?)", (filepath, mtime, file_hash))
    return True

def parse_request_excel(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        # Mapping from prompt: 品名=F10, 依頼No=Q8, 数量=F13, 生地色=F16, 用途=F19, 納期=F22, 出荷先=F31, 規格=F34+L34, 備考=F42
        def g(cell):
            v = ws[cell].value
            return str(v).strip() if v is not None else ""
            
        req_no = g('Q8')
        if not req_no:
            # Fallback to filename
            m = re.search(r'#(\d+)', os.path.basename(filepath))
            if m: req_no = m.group(1)
            
        if not req_no:
            return None # Cannot identify
            
        data = {
            'request_no': req_no,
            'issue_date': "", # Need to extract if available
            'hinmei': g('F10'),
            'qty': g('F13'),
            'kiji': g('F16'),
            'yoto': g('F19'),
            'noki_raw': g('F22'),
            'noki': "",
            'dest': g('F31'),
            'spec': g('F34') + "\n" + g('L34'),
            'biko': g('F42'),
            'prev': "",
            'seiban': "",
            'file': filepath
        }
        return data
    except Exception as e:
        print(f"Error parsing request {filepath}: {e}")
        return None

def parse_bom_excel(filepath):
    # BOM: ファイル名先頭が品番、( )内の #番号 が関連依頼番号、YYMM-NNNN があれば製番
    filename = os.path.basename(filepath)
    # Parse filename
    product_code = filename.split('(')[0].split('.')[0].strip()
    
    ref_reqs = []
    for m in re.finditer(r'#(\d+)', filename):
        ref_reqs.append(m.group(1))
        
    seiban = ""
    sm = re.search(r'(\d{4}-\d{4})', filename)
    if sm: seiban = sm.group(1)
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        components = []
        layout_ok = True
        
        # Determine if it's an exception layout.
        # We expect a horizontal matrix where row 2 or 3 contains keywords like "プレート", "チューブ", "外装", etc.
        keywords = ['プレート', 'チューブ', '外装', '備考', '重量', 'P/K', '役割', '部品']
        header_row_idx = None
        role_cols = {}
        
        for row_idx in range(1, 11):
            row_has_keyword = False
            temp_roles = {}
            for col_idx in range(1, ws.max_column + 1):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
                if val and val != "(            )": # ignore template empty parentheses
                    temp_roles[col_idx] = val
                    if any(kw in val for kw in keywords):
                        row_has_keyword = True
            
            if row_has_keyword and len(temp_roles) > 0:
                header_row_idx = row_idx
                role_cols = temp_roles
                break
                
        if header_row_idx:
            for c_idx, role_name in role_cols.items():
                if any(kw in role_name for kw in ['備考', '重量', 'kg']):
                    continue
                part = str(ws.cell(row=header_row_idx + 1, column=c_idx).value or "").strip()
                
                notes = []
                for r_idx in range(header_row_idx + 2, header_row_idx + 6):
                    n_val = str(ws.cell(row=r_idx, column=c_idx).value or "").strip()
                    if n_val:
                        notes.append(n_val)
                        
                if part or notes:
                    components.append({'role': role_name, 'part_no': part, 'note': " / ".join(notes)})
            
            # Extract red text (notes/warnings usually at the bottom or highlighted)
            red_texts = []
            for row in ws.iter_rows(min_row=header_row_idx + 1):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.font and cell.font.color and cell.font.color.rgb:
                        rgb = str(cell.font.color.rgb)
                        # Check if it's a shade of red
                        if (rgb.startswith('FF') and rgb[2:4] > '80' and rgb[4:6] == '00' and rgb[6:8] == '00') or rgb in ('FFFF0000', 'FF0000'):
                            text_val = str(cell.value).strip()
                            if text_val and text_val not in red_texts:
                                red_texts.append(text_val)
            
            if red_texts:
                components.append({'role': '【特記・赤字】', 'part_no': '', 'note': ' / '.join(red_texts)})

        else:
            layout_ok = False
            # Read all cell values for AI
            raw_text = []
            for row in ws.iter_rows(values_only=True):
                raw_text.append("\t".join(str(v) if v is not None else "" for v in row))
            components = "\n".join(raw_text)
            
        return {
            'product_code': product_code,
            'seiban': seiban,
            'ref_requests': ref_reqs,
            'layout_ok': layout_ok,
            'components': components,
            'file': filepath
        }
    except Exception as e:
        print(f"Error parsing BOM {filepath}: {e}")
        return None

def import_all():
    # openpyxlの画像形式に関する警告を非表示にする
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    
    init_db()
    conn = get_db()
    c = conn.cursor()
    
    # 1. Import Requests
    req_files = glob.glob("data/加工依頼書/*.xlsx")
    print(f"加工依頼書の取り込みを開始します (全 {len(req_files)} 件)...")
    for i, fp in enumerate(req_files, 1):
        if i % 100 == 0:
            print(f"  ... {i}/{len(req_files)} 件処理中")
        if check_file_changed(c, fp):
            req = parse_request_excel(fp)
            if req:
                c.execute('''INSERT OR REPLACE INTO requests 
                             (request_no, issue_date, hinmei, qty, kiji, yoto, noki_raw, noki, dest, spec, biko, prev, seiban, file)
                             VALUES (:request_no, :issue_date, :hinmei, :qty, :kiji, :yoto, :noki_raw, :noki, :dest, :spec, :biko, :prev, :seiban, :file)''', req)
                # Ensure product exists
                pcode = req['hinmei'].split(' ')[0] # naive extraction
                c.execute("INSERT OR IGNORE INTO products (product_code, name, alias) VALUES (?, ?, ?)", (pcode, req['hinmei'], ""))
                
    # 2. Import BOMs
    bom_files = glob.glob("data/BOM/*.xlsx")
    print(f"BOMの取り込みを開始します (全 {len(bom_files)} 件)...")
    for i, fp in enumerate(bom_files, 1):
        if i % 100 == 0:
            print(f"  ... {i}/{len(bom_files)} 件処理中")
        if check_file_changed(c, fp):
            bom = parse_bom_excel(fp)
            if bom:
                if bom['layout_ok']:
                    c.execute("INSERT OR REPLACE INTO boms (product_code, seiban, layout_ok, file) VALUES (?, ?, ?, ?)",
                              (bom['product_code'], bom['seiban'], True, bom['file']))
                    bom_id = c.lastrowid
                    c.execute("DELETE FROM bom_components WHERE bom_id=?", (bom_id,))
                    for comp in bom['components']:
                        c.execute("INSERT INTO bom_components (bom_id, role, part_no, note) VALUES (?, ?, ?, ?)",
                                  (bom_id, comp['role'], comp['part_no'], comp['note']))
                    
                    c.execute("DELETE FROM bom_requests WHERE bom_id=?", (bom_id,))
                    for ref in bom['ref_requests']:
                        c.execute("INSERT INTO bom_requests (bom_id, request_no) VALUES (?, ?)", (bom_id, ref))
                else:
                    c.execute("INSERT OR REPLACE INTO exceptions (file_path, raw_text, status) VALUES (?, ?, 'pending')",
                              (bom['file'], bom['components']))
                    c.execute("INSERT OR IGNORE INTO boms (product_code, seiban, layout_ok, file) VALUES (?, ?, ?, ?)",
                              (bom['product_code'], bom['seiban'], False, bom['file']))
                              
                c.execute("INSERT OR IGNORE INTO products (product_code, name, alias) VALUES (?, ?, ?)", (bom['product_code'], bom['product_code'], ""))

    # 3. Import Inspections
    insp_files = glob.glob("data/検査証/*.*")
    print(f"検査証の取り込みを開始します (全 {len(insp_files)} 件)...")
    for fp in insp_files:
        if check_file_changed(c, fp):
            # Extract key from filename
            fname = os.path.basename(fp)
            link_key = fname.split('.')[0] # Naive
            c.execute("DELETE FROM simple_files WHERE type='inspection' AND file_path=?", (fp,))
            c.execute("INSERT INTO simple_files (type, file_path, link_key) VALUES ('inspection', ?, ?)", (fp, link_key))

    # 4. Import Photos
    photo_files = glob.glob("data/写真/*.*")
    print(f"写真の取り込みを開始します (全 {len(photo_files)} 件)...")
    for fp in photo_files:
        if check_file_changed(c, fp):
            fname = os.path.basename(fp)
            link_key = fname.split('.')[0]
            # Delete old entries for this file to prevent duplicates without UNIQUE
            c.execute("DELETE FROM simple_files WHERE type='photo' AND file_path=?", (fp,))
            c.execute("INSERT INTO simple_files (type, file_path, link_key) VALUES ('photo', ?, ?)", (fp, link_key))

    # 5. Import Drawings (JWW / WWJ)
    import sys
    jww2pdf_dir = os.path.abspath("data/図面/jww2pdf")
    if jww2pdf_dir not in sys.path:
        sys.path.append(jww2pdf_dir)
    try:
        from render_jww2 import convert
    except ImportError as e:
        print(f"Warning: Could not import render_jww2: {e}")
        convert = None

    drawing_files = glob.glob("data/図面/**/*.jww", recursive=True) + glob.glob("data/図面/**/*.wwj", recursive=True)
    print(f"図面の取り込みを開始します (全 {len(drawing_files)} 件)...")
    for i, fp in enumerate(drawing_files, 1):
        if i % 100 == 0:
            print(f"  ... {i}/{len(drawing_files)} 件処理中")
        if check_file_changed(c, fp):
            # First, clean up any existing entries for this file
            c.execute("DELETE FROM simple_files WHERE type='drawing' AND file_path=?", (fp,))
            c.execute("DELETE FROM previews WHERE file_path=?", (fp,))
            
            # Generate preview
            if convert:
                try:
                    rel_path = os.path.relpath(fp, "data/図面")
                    base_name = os.path.splitext(rel_path)[0]
                    preview_png = os.path.join("static/previews/drawings", base_name + ".png")
                    os.makedirs(os.path.dirname(preview_png), exist_ok=True)
                    
                    need_convert = True
                    if os.path.exists(preview_png) and os.path.getmtime(preview_png) >= os.path.getmtime(fp):
                        need_convert = False
                        c.execute("INSERT INTO previews (file_path, preview_path, page_num) VALUES (?, ?, ?)", (fp, preview_png.replace('\\', '/'), 1))

                    if need_convert and convert(fp, None, preview_png, dpi=300, mode="screen"):
                        c.execute("INSERT INTO previews (file_path, preview_path, page_num) VALUES (?, ?, ?)", (fp, preview_png.replace('\\', '/'), 1))
                except Exception as e:
                    print(f"Failed to convert drawing to preview {fp}: {e}")

            # Read binary and extract text
            try:
                with open(fp, 'rb') as f:
                    data = f.read()
                decoded = data.decode('shift_jis', errors='ignore')
                matches = set(re.findall(r'[1-9]F[0-9]{3}[A-Z]*-[0-9]{3,4}[A-Z]*|IF[0-9]{3}[A-Z]*-[0-9]{3,4}[A-Z]*', decoded))
                
                if matches:
                    for match in matches:
                        c.execute("INSERT INTO simple_files (type, file_path, link_key) VALUES ('drawing', ?, ?)", (fp, match))
                else:
                    # Fallback to filename
                    fname = os.path.basename(fp)
                    link_key = fname.split('.')[0]
                    c.execute("INSERT INTO simple_files (type, file_path, link_key) VALUES ('drawing', ?, ?)", (fp, link_key))
            except Exception as e:
                print(f"Failed to process drawing {fp}: {e}")

        if i % 100 == 0:
            conn.commit()
            
    conn.commit()
    conn.close()
    print("Import finished.")

if __name__ == "__main__":
    import_all()
